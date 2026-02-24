#!/usr/bin/env python3
"""Profile an Excel workbook and export summary JSON."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def sample_header(ws: Worksheet, header_row: int, max_cols: int) -> List[str]:
    headers: List[str] = []
    for col in range(1, max_cols + 1):
        value = ws.cell(row=header_row, column=col).value
        headers.append("" if value is None else str(value))
    return headers


def count_formulas(ws: Worksheet, scan_rows: int, scan_cols: int) -> int:
    formulas = 0
    max_row = min(ws.max_row, scan_rows)
    max_col = min(ws.max_column, scan_cols)
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formulas += 1
    return formulas


def profile_workbook(path: Path, header_row: int, scan_rows: int, scan_cols: int) -> Dict[str, Any]:
    wb = load_workbook(path, data_only=False, read_only=False)
    sheets: List[Dict[str, Any]] = []
    for name in wb.sheetnames:
        ws = wb[name]
        max_col = min(ws.max_column, scan_cols)
        sheets.append(
            {
                "sheet_name": name,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "dimension": ws.calculate_dimension(),
                "formula_cells_scanned": count_formulas(ws, scan_rows, scan_cols),
                "header_sample": sample_header(ws, header_row, max_col),
                "merged_ranges_count": len(ws.merged_cells.ranges),
            }
        )
    wb.close()
    return {
        "file": str(path),
        "sheet_count": len(sheets),
        "header_row": header_row,
        "scan_rows": scan_rows,
        "scan_cols": scan_cols,
        "sheets": sheets,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Profile workbook structure and formula density.")
    parser.add_argument("input", type=Path, help="Input .xlsx/.xlsm path")
    parser.add_argument("--out", type=Path, default=None, help="Output JSON path")
    parser.add_argument("--header-row", type=int, default=1, help="Header row index to sample")
    parser.add_argument("--scan-rows", type=int, default=2000, help="Max rows to scan for formulas")
    parser.add_argument("--scan-cols", type=int, default=200, help="Max columns to scan for formulas/header")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    out = args.out or args.input.with_suffix(".profile.json")
    profile = profile_workbook(args.input, args.header_row, args.scan_rows, args.scan_cols)
    out.write_text(json.dumps(profile, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote profile: {out}")


if __name__ == "__main__":
    main()
