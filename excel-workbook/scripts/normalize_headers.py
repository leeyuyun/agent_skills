#!/usr/bin/env python3
"""Normalize worksheet header row to deterministic snake_case names."""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


def to_snake(text: str) -> str:
    text = text.strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "column"


def unique_name(base: str, used: Dict[str, int]) -> str:
    if base not in used:
        used[base] = 1
        return base
    used[base] += 1
    return f"{base}_{used[base]}"


def normalize_headers(input_path: Path, sheet: str, row_index: int, output_path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(input_path)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet}' not found in workbook.")

    ws = wb[sheet]
    changes: List[Dict[str, str]] = []
    used: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_index, column=col)
        original = "" if cell.value is None else str(cell.value)
        normalized = unique_name(to_snake(original), used)
        cell.value = normalized
        changes.append({"column_index": str(col), "from": original, "to": normalized})

    wb.save(output_path)
    wb.close()
    return changes


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Normalize header row to snake_case names.")
    parser.add_argument("input", type=Path, help="Input workbook path")
    parser.add_argument("--sheet", required=True, help="Target worksheet name")
    parser.add_argument("--row", type=int, default=1, help="Header row index")
    parser.add_argument("--out", type=Path, default=None, help="Output workbook path")
    parser.add_argument("--inplace", action="store_true", help="Modify input workbook in place")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.inplace:
        output = args.input
    else:
        output = args.out or args.input.with_name(f"{args.input.stem}.normalized{args.input.suffix}")

    changes = normalize_headers(args.input, args.sheet, args.row, output)
    print(f"Wrote workbook: {output}")
    print("Header changes:")
    for c in changes:
        print(f"  col {c['column_index']}: '{c['from']}' -> '{c['to']}'")


if __name__ == "__main__":
    main()
